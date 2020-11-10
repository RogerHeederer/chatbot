import openpyxl #엑셀 파일을 다루는 라이브러리
import random
from openpyxl import Workbook, load_workbook

def chatbot_data_cleansing(): # 함수 추가 by Roger
    root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
    chatbot_data_raw = root_path + "/chatbot_data.txt"
    chatbot_data_cleaned = root_path + "/chatbot_data_cleaned.txt"
    
    chat_dialog = open(chatbot_data_raw, 'r')
    chat_lines = chat_dialog.readlines()
    
    f = open(chatbot_data_cleaned, 'w')
    
    for line_data in chat_lines:
        line_data = line_data.replace('"',"").replace(".","") #쌍따옴표, 마침표 제거
        f.write(line_data)
    f.close()
    
def chatbot_question_data(): #함수 추가 by Roger
    
    root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
    chatbot_data_raw = root_path + "/chatbot_data_cleaned.txt"
    chatbot_q_data = root_path + "/chatbot_question_data.txt"
    
    chat_dialog = open(chatbot_data_raw, 'r')
    chat_lines = chat_dialog.readlines()

    f = open(chatbot_q_data, 'w')
    
    for line_data in chat_lines:
        data = line_data.split(',')
        # 카테고리 / 질문 구성
        f.write(data[-1].strip() + "    " + data[0].strip() + "\n")
    f.close()
   
def chatbot_answer_data(): #함수 추가 by Roger
    
    root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
    chatbot_data_cleaned = root_path + "/chatbot_data_cleaned.txt"
    chatbot_a_data = root_path + "/chatbot_answer_data.txt"
    
    chat_dialog = open(chatbot_data_cleaned, 'r')
    chat_lines = chat_dialog.readlines()

    category = ['0', '1', '2', 'label']
    f = open(chatbot_a_data, 'w')
    
    # 카테고리/ 답변 or 카테고리/ 답변,답변  형태로 구성
    for line_data in chat_lines:
        data = line_data.split(',')
        if data[2].strip() in category: #공백 제거 해야지 카테고리랑 비교가능
        # data[2]가 카테고리이면 질문/답변/카테고리 형태
            f.write(data[-1].strip() + "    " + data[1].strip() + "\n")
        else: # 그렇지 않으면 질문/답변,답변/카테고리 형태
            f.write(data[-1].strip() + "    " + data[1].strip() + data[2].strip() + "\n")
    f.close()

def chatbot_dialog_for_autoregressive_data(): #함수 추가 by Roger
    
    root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
    chatbot_data_cleaned = root_path + "/chatbot_data_cleaned.txt"
    chatbot_dialog_data = root_path + "/chatbot_dialog_for_autoregressive.txt"
    
    chat_dialog = open(chatbot_data_cleaned, 'r')
    chat_lines = chat_dialog.readlines()

    category = ['0', '1', '2', 'label']
    f = open(chatbot_dialog_data, 'w')
    
    # 카테고리/ 답변 or 카테고리/ 답변,답변  형태로 구성
    for line_data in chat_lines:
        data = line_data.split(',')
        if data[2].strip() in category: #공백 제거 해야지 카테고리랑 비교가능
        # data[2]가 카테고리이면 질문/답변/카테고리 형태
            f.write(data[0].strip() + "    " + data[1].strip() + "\n")
        else: # 그렇지 않으면 질문/답변,답변/카테고리 형태
            f.write(data[0].strip() + "    " + data[1].strip() + data[2].strip() + "\n")
    f.close()

def chatbot_autoregressive_data_with_token(): #함수 추가 by Roger
    
    root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
    chatbot_data_cleaned = root_path + "/chatbot_data_cleaned.txt"
    chatbot_dialog_data_token = root_path + "/chatbot_dialog_for_autoregressive_with_token.txt"
    
    chat_dialog = open(chatbot_data_cleaned, 'r')
    chat_lines = chat_dialog.readlines()

    category = ['0', '1', '2', 'label']
    f = open(chatbot_dialog_data_token, 'w')
    
    # 카테고리/ 답변 or 카테고리/ 답변,답변  형태로 구성
    for line_data in chat_lines:
        data = line_data.split(',')
        if data[2].strip() in category: #공백 제거 해야지 카테고리랑 비교가능
        # data[2]가 카테고리이면 질문/답변/카테고리 형태
            f.write("<s>" + data[0].strip() + "</s><s>" + data[1].strip() + "</s>" + "\n")
        else: # 그렇지 않으면 질문/답변,답변/카테고리 형태
            f.write("<s>" + data[0].strip() + "</s><s>" + data[1].strip() + data[2].strip() + "</s>" + "\n")
    f.close()
    

def wellness_fulldialog_transfer_txt(): # 함수 추가 by Roger
    root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
    wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
    wellness_full_output = root_path + "/wellness_dialog.txt"

    f = open(wellness_full_output, 'w') #파일 오픈
    wb = load_workbook(filename=wellness_file) 
    ws = wb[wb.sheetnames[0]] # 첫번째 시트 세팅

    #위 xlsx 파일은 3개의 컬럼으로 구성됨. 감정구분 / 질문 / 답변
    for row in ws.iter_rows(): # 로우 단위로 작업
        if row[2].value == None: #답변이 없는 경우
            f.write(row[0].value + "    " + row[1].value + "\n")
        else:
            f.write(row[0].value + "    " + row[1].value + "    " + row[2].value + "\n")
    f.close() #파일 닫기    
    
def wellness_question_data(): # 질문만 따로 떼어내서 파일 만들기
    root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
    wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
    wellness_q_output = root_path + "/wellness_dialog_question.txt"

    f = open(wellness_q_output, 'w')
    wb = load_workbook(filename=wellness_file)
    ws = wb[wb.sheetnames[0]]

    for row in ws.iter_rows():
        f.write(row[0].value + "    " + row[1].value + "\n") # 감정구분/tab/질문
    f.close()

def wellness_answer_data(): # 답변만 따로 떼어내서 파일 만들기
    root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
    wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
    wellness_a_output = root_path + "/wellness_dialog_answer.txt"

    f = open(wellness_a_output, 'w')
    wb = load_workbook(filename=wellness_file)
    ws = wb[wb.sheetnames[0]]

    for row in ws.iter_rows():
        if row[2].value == None:
            continue
        else:
            f.write(row[0].value + "    " + row[2].value + "\n") # 감정구분/tab/답변
    f.close()

def wellness_dialog_for_autoregressive():
    # AutoRegressive 언어 모델에 맞는 인풋 만들기, 해당 개념은 따로 찾아보기
    root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
    wellness_answer_file = root_path + "/wellness_dialog_answer.txt"
    wellness_question_file = root_path + "/wellness_dialog_question.txt"
    wellness_autoregressive_file = root_path + "/wellness_dialog_for_autoregressive.txt"

    answ_file = open(wellness_answer_file, 'r')
    ques_file = open(wellness_question_file, 'r')
    autoregressive_file = open(wellness_autoregressive_file, 'w')

    answ_lines = answ_file.readlines()
    ques_lines = ques_file.readlines()
    ques_dict = {}

    for line_num, line_data in enumerate(ques_lines):
        ques_data = line_data.split('    ') #감정구분/질문  split
        for ans_line_num, ans_line_data in enumerate(answ_lines):
            ans_data = ans_line_data.split('    ') #  감정구분/답변 split
            if ques_data[0] == ans_data[0]: #감정구분 값이 같다면
                #질문값 + "    " + 대답값 더한다
                autoregressive_file.write(ques_data[1][:-1] + "    " + ans_data[1])
            else:
                continue
    answ_file.close()
    ques_file.close()
    autoregressive_file.close()

def wellness_autoregressive_data_with_token():
    # 질문과 답변으로 구성하는 autoregressive 데이터 형태로 전처리함과 동시에 <s></s> 토큰 값도 추가해준다.
    root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
    wellness_answer_file = root_path + "/wellness_dialog_answer.txt"
    wellness_question_file = root_path + "/wellness_dialog_question.txt"
    wellness_autoregressive_file = root_path + "/wellness_dialog_for_autoregressive_with_token.txt"

    answ_file = open(wellness_answer_file, 'r')
    ques_file = open(wellness_question_file, 'r')
    autoregressive_file = open(wellness_autoregressive_file, 'w')

    answ_lines = answ_file.readlines()
    ques_lines = ques_file.readlines()
    ques_dict = {}

    for line_num, line_data in enumerate(ques_lines):
        ques_data = line_data.split('    ') #감정구분/질문  split
        for ans_line_num, ans_line_data in enumerate(answ_lines):
            ans_data = ans_line_data.split('    ') #  감정구분/답변 split
            if ques_data[0] == ans_data[0]: #감정구분 값이 같다면
                #질문값 + "    " + 대답값 더한다 + <s>,</s> 시작과 끝을 나타내는 토큰 값을 추가해준다
                autoregressive_file.write("<s>" + ques_data[1][:-1] + "</s><s>" + ans_data[1][:-1] + "</s>\n")
            else:
                continue
    answ_file.close()
    ques_file.close()
    autoregressive_file.close()

def seperate_wellness_data():
    # Autoregressive data를 훈련 세트와 테스트 세트로 나누기

    root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
    file_path = root_path + "/wellness_dialog_for_autoregressive.txt"
    train_file_path = root_path + "/wellness_dialog_for_autoregressive_train.txt"
    test_file_path = root_path + "/wellness_dialog_for_autoregressive_test.txt"

    seperated_file = open(file_path, 'r')
    train_file = open(train_file_path, 'w')
    test_file = open(test_file_path, 'w')

    seperated_file_lines = seperated_file.readlines()
    ques_dict = {}

    for line_num, line_data in enumerate(seperated_file_lines):
        # 90% : 10% 비율로 나누는 방식
        rand_num = random.randint(0, 10) # 랜덤 숫자 0 - 10 부여
        if rand_num < 10: # 0~9 범주면 훈련 파일에 텍스트 작성
            train_file.write(line_data)
        else: # 10이면 테스트 파일에 작성
            test_file.write(line_data)

    seperated_file.close()
    train_file.close()
    test_file.close()
    
def seperate_total_data(): # By Roger
  root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
  file_path = root_path + "/total.txt"
  train_file_path = root_path + "/total_train.txt"
  test_file_path = root_path + "/total_test.txt"

  sperated_file = open(file_path, 'r')
  train_file = open(train_file_path, 'w')
  test_file = open(test_file_path, 'w')

  sperated_file_lines = sperated_file.readlines()
  ques_dict = {}
  for line_num, line_data in enumerate(sperated_file_lines):
    rand_num = random.randint(0, 10)
    if rand_num < 10:
      train_file.write(line_data)
    else:
      test_file.write(line_data)

  sperated_file.close()
  train_file.close()
  test_file.close()

def merge_data():
    # chatbot_data.txt와 wellness_dialog.txt 합치기
    root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
    chatbot_file = root_path + "/chatbot_data.txt"
    wellness_file = root_path + "/wellness_dialog.txt"
    total_data_file = root_path + "/chatbot_wellness_data.txt"

    chatbot_f = open(chatbot_file, 'r')
    wellness_f = open(wellness_file, 'r')
    output_f = open(total_data_file, 'w')

    chatbot_lines = chatbot_f.readlines()
    for line_num, line_data in enumerate(chatbot_lines):
        output_f.write(line_data)
    
    wellness_lines = wellness_f.readlines()
    for line_num, line_data in enumerate(wellness_lines):
        output_f.write(line_data)

    chatbot_f.close()
    wellness_f.close()
    output_f.close()

def merge_autoregressive_data():

  root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
  chatbot_file = root_path + "/chatbot_dialog_for_autoregressive.txt"
  wellness_file = root_path + "/wellness_dialog_for_autoregressive.txt"

  total_data_file = root_path + "/total.txt"

  chatbot_f = open(chatbot_file, 'r')
  wellness_f = open(wellness_file, 'r')
  output_f = open(total_data_file, 'w')

  chatbot_lines = chatbot_f.readlines()
  for line_num, line_data in enumerate(chatbot_lines):
    output_f.write(line_data)

  wellness_lines = wellness_f.readlines()
  for line_num, line_data in enumerate(wellness_lines):
    output_f.write(line_data)

  chatbot_f.close()
  wellness_f.close()
  output_f.close()
  
  
def merge_token_data():

  root_path = "/content/drive/My Drive/RogerHeederer/ChatBot/KoGPT2_Wellness/data"
  chatbot_file = root_path + "/chatbot_dialog_for_autoregressive_with_token.txt"
  wellness_file = root_path + "/wellness_dialog_for_autoregressive_with_token.txt"

  total_data_file = root_path + "/total_token.txt"

  chatbot_f = open(chatbot_file, 'r')
  wellness_f = open(wellness_file, 'r')
  output_f = open(total_data_file, 'w')

  chatbot_lines = chatbot_f.readlines()
  for line_num, line_data in enumerate(chatbot_lines):
    output_f.write(line_data)

  wellness_lines = wellness_f.readlines()
  for line_num, line_data in enumerate(wellness_lines):
    output_f.write(line_data)

  chatbot_f.close()
  wellness_f.close()
  output_f.close()
  


  
if __name__ == "__main__":
    #Chatbot_data.txt 전처리#
    
     chatbot_data_cleansing()
     chatbot_question_data()
     chatbot_answer_data()
     chatbot_dialog_for_autoregressive_data()
     chatbot_autoregressive_data_with_token()
    
    #Wellness_data.xlsx 전처리#

     wellness_fulldialog_transfer_txt()
     wellness_question_data()
     wellness_answer_data()
     wellness_dialog_for_autoregressive()
     wellness_autoregressive_data_with_token()

    #전처리 된 챗봇 데이터와 웰니스 데이터 합치기
     merge_autoregressive_data() # 데이터 로드 인풋용
    # #merge_token_data()
  
    #합쳐진 데이터를 훈련/테스트 셋으로 나누기
     seperate_total_data()